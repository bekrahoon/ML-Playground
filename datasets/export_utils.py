import io
from datetime import datetime
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


def generate_pdf_report(ml_model):
    """Генерация PDF отчета о модели"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Заголовок
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#2C3E50'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    title = Paragraph(f'ML Model Report: {ml_model.name}', title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Информация о модели
    info_style = styles['Normal']
    info_data = [
        ['Параметр', 'Значение'],
        ['Название модели', ml_model.name],
        ['Алгоритм', ml_model.get_algorithm_display()],
        ['Датасет', ml_model.dataset.name],
        ['Целевая переменная', ml_model.target_column],
        ['Дата создания', ml_model.created_at.strftime('%Y-%m-%d %H:%M')],
    ]
    
    info_table = Table(info_data, colWidths=[2.5*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498DB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(info_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Метрики
    metrics_title = Paragraph('<b>Метрики модели</b>', styles['Heading2'])
    elements.append(metrics_title)
    elements.append(Spacer(1, 0.1*inch))
    
    metrics_data = [['Метрика', 'Значение']]
    
    if ml_model.accuracy:
        metrics_data.append(['Accuracy', f'{ml_model.accuracy:.4f}'])
    if ml_model.f1_score:
        metrics_data.append(['F1 Score', f'{ml_model.f1_score:.4f}'])
    if ml_model.mse:
        metrics_data.append(['MSE', f'{ml_model.mse:.4f}'])
    if ml_model.rmse:
        metrics_data.append(['RMSE', f'{ml_model.rmse:.4f}'])
    if ml_model.r2_score:
        metrics_data.append(['R² Score', f'{ml_model.r2_score:.4f}'])
    
    metrics_table = Table(metrics_data, colWidths=[3*inch, 3*inch])
    metrics_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2ECC71')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(metrics_table)
    
    # Confusion Matrix (если есть)
    if ml_model.confusion_matrix:
        elements.append(Spacer(1, 0.3*inch))
        cm_title = Paragraph('<b>Confusion Matrix</b>', styles['Heading2'])
        elements.append(cm_title)
        elements.append(Spacer(1, 0.1*inch))
        
        cm_data = [['', 'Predicted 0', 'Predicted 1']]
        cm = ml_model.confusion_matrix
        cm_data.append(['Actual 0', str(cm[0][0]), str(cm[0][1])])
        cm_data.append(['Actual 1', str(cm[1][0]), str(cm[1][1])])
        
        cm_table = Table(cm_data, colWidths=[2*inch, 2*inch, 2*inch])
        cm_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E74C3C')),
            ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#E74C3C')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('TEXTCOLOR', (0, 1), (0, -1), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(cm_table)
    
    # Описание
    if ml_model.description:
        elements.append(Spacer(1, 0.3*inch))
        desc_title = Paragraph('<b>Описание</b>', styles['Heading2'])
        elements.append(desc_title)
        elements.append(Spacer(1, 0.1*inch))
        desc = Paragraph(ml_model.description, styles['Normal'])
        elements.append(desc)
    
    # Генерация PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_excel_report(ml_model):
    """Генерация Excel отчета о модели"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Model Report"
    
    # Стили
    header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    metrics_fill = PatternFill(start_color="2ECC71", end_color="2ECC71", fill_type="solid")
    
    # Заголовок
    ws['A1'] = 'ML Model Report'
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:B1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Информация о модели
    ws['A3'] = 'Параметр'
    ws['B3'] = 'Значение'
    ws['A3'].fill = header_fill
    ws['B3'].fill = header_fill
    ws['A3'].font = header_font
    ws['B3'].font = header_font
    
    row = 4
    info = [
        ('Название модели', ml_model.name),
        ('Алгоритм', ml_model.get_algorithm_display()),
        ('Датасет', ml_model.dataset.name),
        ('Целевая переменная', ml_model.target_column),
        ('Дата создания', ml_model.created_at.strftime('%Y-%m-%d %H:%M')),
    ]
    
    for param, value in info:
        ws[f'A{row}'] = param
        ws[f'B{row}'] = value
        row += 1
    
    # Метрики
    row += 1
    ws[f'A{row}'] = 'Метрика'
    ws[f'B{row}'] = 'Значение'
    ws[f'A{row}'].fill = metrics_fill
    ws[f'B{row}'].fill = metrics_fill
    ws[f'A{row}'].font = header_font
    ws[f'B{row}'].font = header_font
    
    row += 1
    metrics = []
    if ml_model.accuracy:
        metrics.append(('Accuracy', f'{ml_model.accuracy:.4f}'))
    if ml_model.f1_score:
        metrics.append(('F1 Score', f'{ml_model.f1_score:.4f}'))
    if ml_model.mse:
        metrics.append(('MSE', f'{ml_model.mse:.4f}'))
    if ml_model.rmse:
        metrics.append(('RMSE', f'{ml_model.rmse:.4f}'))
    if ml_model.r2_score:
        metrics.append(('R² Score', f'{ml_model.r2_score:.4f}'))
    
    for metric, value in metrics:
        ws[f'A{row}'] = metric
        ws[f'B{row}'] = value
        row += 1
    
    # Confusion Matrix
    if ml_model.confusion_matrix:
        row += 1
        ws[f'A{row}'] = 'Confusion Matrix'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        cm = ml_model.confusion_matrix
        ws[f'B{row}'] = 'Predicted 0'
        ws[f'C{row}'] = 'Predicted 1'
        row += 1
        ws[f'A{row}'] = 'Actual 0'
        ws[f'B{row}'] = cm[0][0]
        ws[f'C{row}'] = cm[0][1]
        row += 1
        ws[f'A{row}'] = 'Actual 1'
        ws[f'B{row}'] = cm[1][0]
        ws[f'C{row}'] = cm[1][1]
    
    # Автоширина колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохранение в буфер
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer